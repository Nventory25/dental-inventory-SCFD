import streamlit as st
import sqlite3
import pandas as pd
import string
import re
import hashlib
import os
import time





from datetime import datetime, date

#EMAIL SENT TO INBOX FOR FORGETTON PASSWORD
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_reset_email(to_email, reset_code):
    gmail_user = 'myinventorytool@gmail.com'  # ✅ Your Gmail
    app_password = 'fnjdppvphjelcygo'  # ✅ App password (no quotes or spaces)

    subject = '🔐 Password Reset Code'
    body = f'Your reset code is: {reset_code}'

    msg = MIMEMultipart()
    msg['From'] = gmail_user
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, app_password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print("Email failed:", e)
        return False



# === SET PAGE CONFIG ===

st.set_page_config(page_title="Dental Inventory App", layout="wide")





# === HIDE STREAMLIT BRANDING ===

st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)



def trigger_action(key, delta):
    st.session_state[key] = max(0, st.session_state.get(key, 0) + delta)


DB_PATH = "inventory.db"





def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS products (
            name TEXT PRIMARY KEY,
            part_number TEXT,
            target_qty INTEGER,
            current_qty INTEGER,
            supplier_link TEXT,
            price_per_unit REAL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS restock_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user_name TEXT,
            product TEXT,
            qty_added INTEGER
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS purchase_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user_name TEXT,
            product_name TEXT,
            qty_purchased INTEGER
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_name TEXT,
            product_name TEXT,
            qty_removed INTEGER,
            timestamp TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            name TEXT PRIMARY KEY,
            password TEXT,
            email TEXT
        )
    """)

    conn.commit()
    conn.close()

if not os.path.exists(DB_PATH):
    print("🛠️ No database found — creating new one.")
    init_db()

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    
if 'user_name' not in st.session_state:
    st.session_state.user_name = ""


if not st.session_state.logged_in:
    # all of your login, forgot password, and register forms go here





# === TITLE AND LOGO CENTERED ===

    import base64

    with open("logo.png", "rb") as image_file:
        encoded = base64.b64encode(image_file.read()).decode()

    st.markdown(
        f"""
        <div style='text-align: center;'>
            <img src='data:image/png;base64,{encoded}' width='60' style='margin-bottom: 10px;' />
            <h1>Nventory</h1>
        </div>
        """,
        unsafe_allow_html=True
    )







    # === LOGIN / REGISTER ===
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "user_name" not in st.session_state:
        st.session_state.user_name = ""
    if "password_reset_success" not in st.session_state:
        st.session_state["password_reset_success"] = False

    def hash_password(password):
        return hashlib.sha256(password.encode()).hexdigest()

    def check_credentials(username, password):
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("SELECT password FROM users WHERE name = ?", (username,))
        result = cursor.fetchone()
        conn.close()
        if result:
            return hash_password(password) == result[0]
        return False

    if not st.session_state.logged_in:
        #st.title("Inventory App")

        # Check for first-time setup
        conn = sqlite3.connect("inventory.db")
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        user_count = cursor.fetchone()[0]
        conn.close()

        if user_count == 0:
            st.subheader("👋 First-Time Setup")
            new_user = st.text_input("Create Username")
            new_pass = st.text_input("Create Password", type="password")
            new_email = st.text_input("Email Address")

            if st.button("Create Admin Account"):
                if new_user and new_pass and new_email:
                    conn = sqlite3.connect("inventory.db")
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO users (name, password, email) VALUES (?, ?, ?)",
                                   (new_user, hash_password(new_pass), new_email))
                    conn.commit()
                    conn.close()
                    st.session_state.logged_in = True
                    st.session_state.user_name = new_user
                    st.success("✅ Admin account created! Logging you in...")
                    st.rerun()
                else:
                    st.error("Please enter all fields.")
        else:
            st.subheader("🔐 Login")

            # --- Username Dropdown with Optional Manual Entry ---
            # --- Username Dropdown (Only from DB) ---
            conn = sqlite3.connect("inventory.db")
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM users")
            user_list = [row[0] for row in cursor.fetchall()]
            conn.close()

            selected_option = st.selectbox("Select your username", user_list)
            username = selected_option  # ✅ This is critical!

            password = st.text_input("Password", type="password")
            remember_me = False  # ✅ You can change this if you want real "Remember Me" functionality




            if st.button("Login"):
                if username and password:
                    if check_credentials(username, password):
                        st.session_state.logged_in = True
                        st.session_state.user_name = username
                        if remember_me:
                            st.session_state["remember_me"] = True
                        st.success(f"✅ Welcome back, {username}!")
                        st.rerun()
                    else:
                        st.error("❌ Incorrect username or password.")
                else:
                    st.error("Please enter your username and password.")

    st.markdown("---")


    #FORGOT PASSWORD

    with st.expander("🔁 Forgot Password?"):
        st.subheader("Reset Your Password")
        reset_username = st.text_input("Enter your username", key="reset_username_input")
        if st.button("Send Reset Code"):
            import random
            reset_code = str(random.randint(100000, 999999))

            # Look up email from username
            conn = sqlite3.connect("inventory.db")
            cursor = conn.cursor()
            cursor.execute("SELECT email FROM users WHERE name = ?", (reset_username,))
            result = cursor.fetchone()
            conn.close()

            if result:
                email_to_send = result[0]
                print("Sending reset code to:", email_to_send)  # Debug
                email_success = send_reset_email(email_to_send, reset_code)

                if email_success:
                    st.success("📬 Reset code sent to the email on file.")
                    st.session_state["reset_email"] = email_to_send
                    st.session_state["reset_code"] = reset_code
                    st.session_state["reset_username"] = reset_username
                else:
                    st.error("❌ Email sending failed. Check Gmail credentials.")
            else:
                st.error("❌ Username not found.")


        # 🔁 Password Reset Code Verification





        # ✅ Show success message after password reset
        if st.session_state.get("password_reset_success"):
            st.success("✅ Password successfully reset! Please log in")
            # Optionally clear the success state after showing
            st.session_state["password_reset_success"] = False

        # 🔁 Show reset form if reset_code is in session and password wasn't reset yet
        elif "reset_code" in st.session_state and not st.session_state.get("password_reset_success"):
            st.markdown("---")
            st.subheader("🔑 Enter Reset Code and New Password")

            entered_code = st.text_input("Enter the reset code sent to your email", key="code_input")
            new_password = st.text_input("Enter your new password", type="password", key="new_pw_input")
            confirm_password = st.text_input("Confirm your new password", type="password", key="confirm_pw_input")

            if st.button("Reset Password"):
                if entered_code == st.session_state["reset_code"]:
                    if new_password == confirm_password and new_password:
                        # ✅ Update password in database
                        conn = sqlite3.connect("inventory.db")
                        cursor = conn.cursor()
                        hashed_pw = hash_password(new_password)
                        cursor.execute("UPDATE users SET password = ? WHERE name = ?",
                                       (hashed_pw, st.session_state["reset_username"]))
                        conn.commit()
                        conn.close()

                        # ✅ Clear all session reset info
                        del st.session_state["reset_code"]
                        del st.session_state["reset_email"]
                        del st.session_state["reset_username"]

                        # ✅ Set success flag
                        st.session_state["password_reset_success"] = True
                        st.rerun()
                    else:
                        st.error("❌ Passwords do not match or are empty.")
                else:
                    st.error("❌ Incorrect reset code.")


                       


    # 👇 Optional Features
    st.markdown("---")
    with st.expander("🆕 New User? Register here"):
        reg_email = st.text_input("Email")
        reg_user = st.text_input("Choose a username")
        reg_pass = st.text_input("Choose a password", type="password")

        if st.button("Register"):
            if reg_user and reg_pass and reg_email:
                conn = sqlite3.connect("inventory.db")
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM users WHERE name = ?", (reg_user,))
                if cursor.fetchone():
                    st.error("❌ Username already exists.")
                else:
                    cursor.execute("INSERT INTO users (name, password, email) VALUES (?, ?, ?)",
                                    (reg_user, hash_password(reg_pass), reg_email))
                    conn.commit()
                    st.success("✅ User registered successfully!")
                conn.close()
            else:
                st.error("Please fill out all fields.")


    st.stop()




#OLD LOG OUT BUTTON

#if st.button("🔓 Logout"):
    #st.session_state.logged_in = False
    #st.session_state.user_name = ""
    #st.rerun()





# Reset manually typed purchase input
if "just_purchased_reset" in st.session_state:
    for key in st.session_state["just_purchased_reset"]:
        st.session_state[key] = 0
    del st.session_state["just_purchased_reset"]





from datetime import datetime

def log_purchase(product_name, part_number, quantity, price_per_unit, total_price, user_name):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        INSERT INTO purchase_logs (
            timestamp,
            user_name,
            product_name,
            qty_purchased,
            part_number,
            quantity,
            price_per_unit,
            total_price
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        timestamp,
        user_name,
        product_name,
        quantity,
        part_number,
        quantity,
        price_per_unit,
        total_price
    ))

    conn.commit()
    conn.close()







def get_products():
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT name, part_number, target_qty, current_qty, supplier_link, price_per_unit
        FROM products
        ORDER BY name COLLATE NOCASE
    """)
    products = cursor.fetchall()
    conn.close()
    return products


def get_users():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM users ORDER BY name")
    users = [row[0] for row in cursor.fetchall()]
    conn.close()
    return users

def update_inventory(product_name, qty_removed):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE products
        SET current_qty = current_qty - ?
        WHERE name = ?
    """, (qty_removed, product_name))
    conn.commit()
    conn.close()

def log_action(user_name, product_name, qty_removed):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("""
        INSERT INTO logs (user_name, product_name, qty_removed, timestamp)
        VALUES (?, ?, ?, ?)
    """, (user_name, product_name, qty_removed, timestamp))
    conn.commit()
    conn.close()



# === SIDEBAR BRANDING ===
import base64

def get_base64_logo(logo_path):
    with open(logo_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# === SIDEBAR BRANDING ===
if st.session_state.logged_in:
    with st.sidebar:
        logo_base64 = get_base64_logo("logo.png")
        
        st.markdown(f"""
            <div style='text-align: center;'>
                <img src='data:image/png;base64,{logo_base64}' width='30' style='margin-bottom: 10px;'/>
                <h4 style='margin-bottom: 4px;'>Nventory</h4>
                <small>Smart Inventory for Everyone</small>
            </div>
        """, unsafe_allow_html=True)





# === SAFE RESET HANDLER ===
if "just_submitted_qty_reset" in st.session_state:
    reset_key = st.session_state.just_submitted_qty_reset
    if reset_key in st.session_state:
        st.session_state[reset_key] = 0
    del st.session_state.just_submitted_qty_reset






# Handle reset from last confirmed restock
if "just_submitted_restock_reset" in st.session_state:
    reset_key = st.session_state.just_submitted_restock_reset
    if reset_key in st.session_state:
        st.session_state[reset_key] = 0
    del st.session_state.just_submitted_restock_reset





# === UI ===
st.set_page_config(page_title="Dental Inventory App", layout="wide")
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Inventory", "Add/Edit Product", "Manage Users", "Reports", "Admin Upload"])






# === Handle purchase logging on button click ===
if "pending_purchase" in st.session_state:
    log_data = st.session_state.pop("pending_purchase")
    log_purchase(**log_data)
    st.success(f"🛒 Logged purchase of {log_data['quantity']} × {log_data['product_name']}")





#NEW SIDEBAR DROPDOWN
with st.sidebar:
    st.markdown("### 👤 Account")
    st.write(f"Logged in as: **{st.session_state['user_name']}**")

    # Fetch all usernames from DB
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM users")
    user_list = [row[0] for row in cursor.fetchall()]
    conn.close()

    # Set the active user to whoever logged in by default
    if "active_user" not in st.session_state:
        st.session_state["active_user"] = st.session_state["user_name"]
        #current_user = st.session_state["active_user"]

    selected_user = st.selectbox(
        "Log actions as:",
        user_list,
        index=user_list.index(st.session_state["active_user"]),
        key="active_user_select"
    )
    st.session_state["active_user"] = selected_user

    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()







# === PAGE: Inventory ===
if page == "Inventory":
    st.title("📦 Inventory Management")

    if st.session_state.get("just_submitted", False):
        for key in list(st.session_state.keys()):
            if key.startswith("qty_") or key.startswith("input_") or key.startswith("manual_input_") or key.startswith("restock_input_") or key.startswith("purchase_"):
                st.session_state[key] = 0
        st.session_state.just_submitted = False

    search_query = st.text_input("Search Inventory")
    show_low_only = st.checkbox("Show Only Low Stock", value=False)

    all_products = get_products()





    # === LETTER FILTER BAR ===


    import streamlit as st
    import string

    # === SESSION STATE ===
    if "selected_letter" not in st.session_state:
        st.session_state.selected_letter = "ALL"

    selected_letter = st.session_state.selected_letter

    # === CSS STYLING ===
    st.markdown("""
        <style>
        /* Remove Streamlit's default gray border and fully customize buttons */
        div[data-testid="column"] button {
            border: 1px solid white !important;       /* 🔁 white border */
            background-color: white !important;       /* white background */
            color: black !important;                  /* text color */
            font-weight: bold !important;
            border-radius: 10px !important;
            padding: 6px 14px !important;
            margin: 2px !important;
            box-shadow: none !important;
            transition: all 0.2s ease-in-out !important;
        }

        /* Hover effect */
        div[data-testid="column"] button:hover {
            border: 1px solid #10b981 !important;      /* subtle green */
            background-color: #f0f0f0 !important;
            color: black !important;
        }

        /* Active (clicked) button */
        div[data-testid="column"] button:focus {
            outline: none !important;
            border: 2px solid #10b981 !important;
            background-color: #e6fffa !important;
            color: #10b981 !important;
        }
        </style>
    """, unsafe_allow_html=True)



    # === LETTER FILTER BAR ===
    # === LETTER FILTER BAR (Split into two rows) ===

    selected_letter = st.session_state.get("selected_letter", "ALL")


    # First row: '#' + A–Z
    letters_row1 = ['#'] + list(string.ascii_uppercase)
    cols1 = st.columns(len(letters_row1))

    for idx, letter in enumerate(letters_row1):
        display_label = '&#35;' if letter == '#' else letter
        if cols1[idx].button(display_label, key=f"letter_{letter}"):
            st.session_state.selected_letter = letter
            selected_letter = letter
            st.rerun()

    # === SECOND ROW: "ALL" button + Selected letter text ===
    col_all, col_text = st.columns([1, 4])  # Adjust width ratio as needed

    with col_all:
        if st.button("ALL", key="letter_ALL"):
            st.session_state.selected_letter = "ALL"
            selected_letter = "ALL"
            st.rerun()

    with col_text:
        st.markdown(
            f"<div style='margin-left: -10px; padding-top: 10px;'>Selected letter: <code>{selected_letter}</code></div>",
            unsafe_allow_html=True
        )





    # === END OF FILTER BAR SECTION ===




    #selected_letter = st.session_state.selected_letter


    def starts_with(product_name, letter):
        if letter == "ALL":
            return True
        if letter == "#":
            return not product_name[0].isalpha()
        return product_name.upper().startswith(letter)

    products = [p for p in all_products if starts_with(p[0], selected_letter)]


    for name, part_number, target, current, supplier_link, price in products:
        if show_low_only and current >= 0.65 * target:
            continue
        if search_query.lower() not in name.lower():
            continue

        safe_key = re.sub(r'\W+', '_', name)
        qty_key = f"qty_{safe_key}"
        input_key = f"input_{safe_key}"
        manual_input_key = f"manual_input_{safe_key}"
        restock_input_key = f"restock_input_{safe_key}"
        purchase_qty_key = f"purchase_qty_{safe_key}"
        purchase_input_key = f"purchase_input_{safe_key}"

        for key in [qty_key, input_key, manual_input_key, restock_input_key, purchase_qty_key, purchase_input_key]:
            if key not in st.session_state:
                st.session_state[key] = 0

        is_low = current < 0.65 * target
        bg_style = """
            background-color: #ff4c4c;
            border-left: 6px solid #ff0000;
            padding: 10px;
            border-radius: 8px;
            color: white;
        """ if is_low else """
            background-color: #f7f7f7;
            padding: 10px;
            border-radius: 8px;
        """



        st.markdown("<div style='margin-bottom: 16px;'></div>", unsafe_allow_html=True)



        # === HEADER BOX (top row only, with LOW STOCK highlight and warning) ===
        header_color = "#fee2e2" if is_low else "#f1f5f9"  # light red if low, light grey otherwise
        border_color = "#dc2626" if is_low else "#94a3b8"  # red if low, slate if normal

        st.markdown(f"""
        <div style='border: 2px solid {border_color}; border-radius: 6px; padding: 10px 15px; margin-bottom: 0px; background-color: {header_color};'>
            <div style='display: flex; justify-content: space-between; align-items: center;'>
                <div><strong style='font-size: 1.1rem;'>{name}</strong></div>
                <div style='font-size: 0.9rem; color: #666;'>🔢 <strong>Part #:</strong> {part_number}</div>
                <div style='font-size: 0.9rem; color: #666;'>🎯 <strong>Target Inventory:</strong> {target}</div>
                <div style='font-size: 0.9rem; color: #666;'>📦 <strong>Current Inventory:</strong> {current}</div>
            </div>
            {"<div style='color: #DC2626; font-weight: bold; margin-top: 8px;'>⚠️ LOW STOCK</div>" if is_low else ""}
        </div>
        """, unsafe_allow_html=True)



        #st.markdown("<div style='margin-bottom: 16px;'></div>", unsafe_allow_html=True)

        st.markdown("<div style='padding: 14px 0;'></div>", unsafe_allow_html=True)





        # === INVENTORY TRACKER SECTION ===
        cols = st.columns([8, 1, 1, 1, 1, 2, 3])

        with cols[0]:
            st.markdown("**Inventory Tracker**")

        # Initialize qty_key safely (only once at the top of your loop, outside this block ideally)
        if qty_key not in st.session_state:
            st.session_state[qty_key] = 0

        # ➖ button
        with cols[3]:
            if st.button("➖", key=f"minus_{safe_key}"):
                st.session_state[qty_key] = max(0, st.session_state[qty_key] - 1)
                st.rerun()  # 🚀 Force rerun after update


        # ➕ button
        with cols[5]:
            if st.button("➕", key=f"plus_{safe_key}"):
                st.session_state[qty_key] += 1
                st.rerun()  # 🚀 Force rerun after update


        # Qty input — directly tied to session_state key
        with cols[4]:
            st.number_input(
                "Qty",
                min_value=0,
                step=1,
                key=qty_key,
                label_visibility="collapsed"
            )

        # ✅ Confirm inventory pull
        if cols[6].button("✅ Confirm", key=f"confirm_{safe_key}"):
            qty = st.session_state.get(qty_key, 0)  # ✅ this is the fix!
            if not selected_user:
                st.warning("Please select a user.")
            elif qty <= 0:
                st.warning("Please enter a valid quantity.")
            else:
                update_inventory(name, qty)
                log_action(selected_user, name, qty)
                st.success(f"{qty} removed by {selected_user} for {name}")
                st.session_state.just_submitted_qty_reset = qty_key
                st.rerun()


        st.markdown("<div style='margin-bottom: 30px;'></div>", unsafe_allow_html=True)






        # === PURCHASE SECTION ===
        # === PURCHASE SECTION ===
        # === PURCHASE SECTION ===
        cols = st.columns([8, 1, 1, 1, 1, 2, 3])

        # ✅ 1. Apply pending adjustment BEFORE rendering
        if "pending_purchase_adjust" in st.session_state:
            st.session_state[purchase_qty_key] = max(0, st.session_state[purchase_qty_key] + st.session_state["pending_purchase_adjust"])
            st.session_state[purchase_input_key] = st.session_state[purchase_qty_key]
            del st.session_state["pending_purchase_adjust"]

        # ✅ 2. Reset qty if just purchased
        if "reset_purchase_keys" in st.session_state:
            for key in st.session_state["reset_purchase_keys"]:
                st.session_state[key] = 0
            del st.session_state["reset_purchase_keys"]

        # ✅ 3. Initialize default state
        if purchase_qty_key not in st.session_state:
            st.session_state[purchase_qty_key] = 0
        if purchase_input_key not in st.session_state:
            st.session_state[purchase_input_key] = st.session_state[purchase_qty_key]

        # ✅ 4. Info section: Price and Visit Supplier link
        unit_price = price if price is not None else 0.00
        # ✅ Fix the link if user forgot to add 'https://'
        safe_link = supplier_link if supplier_link.startswith("http") else f"https://{supplier_link}"

        if supplier_link:
            button_html = f"""
                <a href="{supplier_link if supplier_link.startswith('http') else 'https://' + supplier_link}" target="_blank" rel="noopener noreferrer" style="text-decoration: none;">
                    <span style="
                        display: inline-flex;
                        align-items: center;
                        padding: 6px 12px;
                        background-color: #2563eb;
                        color: white;
                        border-radius: 6px;
                        font-size: 0.8rem;
                        font-weight: 500;
                        gap: 6px;
                    ">
                        🔗 Visit Supplier
                    </span>
                </a>
            """
        else:
            button_html = """
                <span style="
                    display: inline-flex;
                    align-items: center;
                    padding: 6px 12px;
                    background-color: #d1d5db;
                    color: #6b7280;
                    border-radius: 6px;
                    font-size: 0.8rem;
                    font-weight: 500;
                    gap: 6px;
                ">
                    🔗 No Link Available
                </span>
            """

        # Inside the layout
        with cols[0]:
            st.markdown(f"""
                <div style="display: flex; align-items: center; gap: 16px; font-size: 0.95rem;">
                    <span style="font-weight: bold;">Purchase</span>
                    <span style="color: green; font-weight: 600;">${unit_price:.2f} / unit</span>
                    {button_html}
                </div>
            """, unsafe_allow_html=True)




        # ✅ 5. ➖ and ➕ buttons (set flag and rerun)
        with cols[3]:
            if st.button("➖", key=f"purchase_minus_{safe_key}"):
                st.session_state["pending_purchase_adjust"] = -1
                st.rerun()

        with cols[5]:
            if st.button("➕", key=f"purchase_plus_{safe_key}"):
                st.session_state["pending_purchase_adjust"] = 1
                st.rerun()

        # ✅ 6. Quantity input field
        with cols[4]:
            st.number_input(
                "", min_value=0, step=1,
                key=purchase_input_key,
                label_visibility="collapsed"
            )

        # ✅ 7. Keep qty in sync
        st.session_state[purchase_qty_key] = st.session_state[purchase_input_key]

        # ✅ 8. Purchase Button
        with cols[6]:
            purchase_qty = st.session_state.get(purchase_input_key, 0)
            if supplier_link:
                purchase_clicked = st.button("🛒 Purchase", key=f"purchase_btn_{safe_key}")
                if purchase_clicked:
                    price_safe = price if price is not None else 0.0
                    total_price = round(price_safe * purchase_qty, 2)

                    log_purchase(
                        product_name=name,
                        part_number=part_number,
                        quantity=purchase_qty,
                        price_per_unit=price_safe,
                        total_price=total_price,
                        user_name=st.session_state.get("current_user", "Unknown")
                    )

                    # ✅ Set keys to reset
                    st.session_state["reset_purchase_keys"] = [purchase_qty_key, purchase_input_key]
                    st.rerun()
            else:
                st.markdown("<span style='color: red;'>No supplier link</span>", unsafe_allow_html=True)



        # ✅ 7. Show supplier link if available
        if "show_supplier_link" in st.session_state:
            supplier_url = st.session_state.pop("show_supplier_link")
            st.markdown(f"""
                <a href="{supplier_url}" target="_blank" rel="noopener noreferrer">
                     <button style="margin-top: 8px; padding:6px 16px; font-weight:bold; background:#10b981; color:white; border:none; border-radius:6px; cursor:pointer;">
                        Visit Supplier 🔗
                    </button>
                </a>
            """, unsafe_allow_html=True)



        st.markdown("<div style='margin-bottom: 30px;'></div>", unsafe_allow_html=True)




        # === THE START OF THE RESTOCK SECTION ===


        # ✅ Initialize restock key
        if restock_input_key not in st.session_state:
            st.session_state[restock_input_key] = 0

        # === RESTOCK SECTION ===
        cols = st.columns([8, 1, 1, 1, 1, 2, 3])

        with cols[0]:
            st.markdown("**Restock**")

        # ➖ button
        with cols[3]:
            if st.button("➖", key=f"restock_minus_{safe_key}"):
                st.session_state[restock_input_key] = max(0, st.session_state[restock_input_key] - 1)
                st.rerun()  # 🚀 rerun after changing state

        # ➕ button
        with cols[5]:
            if st.button("➕", key=f"restock_plus_{safe_key}"):
                st.session_state[restock_input_key] += 1
                st.rerun()  # 🚀 rerun after changing state

        # ⬅️ Number input (must come AFTER button logic!)
        with cols[4]:
            st.number_input(
                "",  # No label
                min_value=0,
                step=1,
                key=restock_input_key,
                label_visibility="collapsed"
            )



        # ✅ Confirm Restock button
        with cols[6]:
            if st.button("✅ Confirm", key=f"restock_confirm_{safe_key}"):
                restock_qty = st.session_state.get(restock_input_key, 0)
                if restock_qty <= 0:
                    st.warning("Enter a valid restock quantity.")
                else:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute("UPDATE products SET current_qty = current_qty + ? WHERE name = ?", (restock_qty, name))
                    conn.commit()
                    conn.close()
                    st.success(f"{restock_qty} restocked for {name}")
                    st.session_state.just_submitted_restock_reset = restock_input_key
                    st.rerun()


        st.markdown("<div style='padding: 12px 0;'></div>", unsafe_allow_html=True)















                

# === PAGE: Add / Edit Product ===
if page == "Add/Edit Product":
    # All the Add Product / Edit Product UI code here

    import streamlit as st
    import sqlite3


    # === DB PATH ===
    DB_PATH = "inventory.db"

    # === GET PRODUCTS FUNCTION ===
    def get_products():
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT name, part_number, target_qty, current_qty, supplier_link, price_per_unit FROM products")
        products = cursor.fetchall()
        conn.close()
        return products


    # === INIT SESSION STATE ===
    if "product_mode" not in st.session_state:
        st.session_state.product_mode = "Edit"
    if "show_success" not in st.session_state:
        st.session_state.show_success = False
    if "clear_fields" not in st.session_state:
        st.session_state.clear_fields = False

    # === PAGE TITLE ===
    st.title("📦 Add or Edit Products")

    # === MODE SWITCH ===
    mode = st.radio("Choose Mode", ["Edit Existing Product", "Add New Product"])
    st.session_state.product_mode = "Add" if mode == "Add New Product" else "Edit"

    # === FETCH PRODUCTS ===
    products = get_products()
    product_names = [p[0] for p in products]

    # Sort from numbers → A–Z
    def sort_key(name):
        is_alpha = name[0].isalpha()
        return (is_alpha, name.lower())

    product_names.sort(key=sort_key)

    selected_product = ""
    if st.session_state.product_mode == "Edit":
        selected_product = st.selectbox("Select a product to edit", product_names)

    # === SHOW SUCCESS MESSAGE ===
    # === SHOW SUCCESS AND FOLLOW-UP OPTIONS ===
    # === SHOW SUCCESS AND FOLLOW-UP BUTTON (AND STOP IMMEDIATELY) ===
    if "add_another_triggered" not in st.session_state:
        st.session_state.add_another_triggered = False
    if "show_success" not in st.session_state:
        st.session_state.show_success = False
    if "show_next_action" not in st.session_state:
        st.session_state.show_next_action = False


    if "delete_timestamp" not in st.session_state:
        st.session_state.delete_timestamp = None
        
    if "product_deleted" not in st.session_state:
        st.session_state.product_deleted = False


    # === SHOW SUCCESS AND FOLLOW-UP BUTTON (AND STOP IMMEDIATELY) ===
    # === SHOW SUCCESS AND "Add Another Product" Button ===
    if st.session_state.get("show_success") and not st.session_state.get("show_next_action"):
        col1, col2 = st.columns([4, 2])

        with col1:
            st.success("✅ Product saved successfully!")

        with col2:
            if st.button("➕ Add/Edit Another Product"):
                st.session_state.clear_fields = True
                st.session_state.show_success = False
                st.session_state.show_next_action = False
                st.rerun()

        st.session_state.last_mode = st.session_state.product_mode
        st.session_state.show_next_action = True
        st.stop()






    # === DEFAULT VALUES ===
    defaults = {
        "name": "",
        "part": "",
        "target": 0,
        "current": 0,
        "supplier": "",
        "price": 0.0,
    }

    if st.session_state.product_mode == "Edit" and selected_product:
        for p in products:
            if p[0] == selected_product:
                defaults = {
                    "name": p[0],
                    "part": p[1],
                    "target": p[2],
                    "current": p[3],
                    "supplier": p[4] or "",
                    "price": float(p[5]) if p[5] is not None else 0.0,
                }

    # === CLEAR FIELDS IF FLAGGED ===
    if st.session_state.clear_fields:
        defaults = {
            k: "" if isinstance(v, str) else 0.0 if k == "price" else 0
            for k, v in defaults.items()
        }

        st.session_state.clear_fields = False

    # === FORM ===
    with st.form("product_form"):
        name = st.text_input("Product Name", value=defaults["name"])
        part_number = st.text_input("Part Number", value=defaults["part"])
        supplier_link = st.text_input("Supplier Link (optional)", value=defaults["supplier"])
        target = st.number_input("6 Month Target Quantity", min_value=0, value=defaults["target"])
        current = st.number_input("Current Stock Quantity", min_value=0, value=defaults["current"])
        price = st.number_input("Price Per Unit ($)", min_value=0.0, step=0.01, format="%.2f", value=defaults["price"])
        submitted = st.form_submit_button("Save Product")

    # === SAVE HANDLER ===
    # === HANDLE SAVE BUTTON SUBMISSION ===
    # === HANDLE SAVE BUTTON SUBMISSION ===
    if submitted:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        if st.session_state.product_mode == "Edit" and selected_product:
            # === Update Existing Product ===
            cursor.execute("""
                UPDATE products
                SET name = ?, part_number = ?, target_qty = ?, current_qty = ?, supplier_link = ?, price_per_unit = ?
                WHERE name = ?
            """, (
                name.strip(), part_number.strip(), target, current,
                supplier_link.strip(), price, selected_product
            ))
        else:
            # === Insert New Product ===
            cursor.execute("""
                INSERT OR IGNORE INTO products (name, part_number, target_qty, current_qty, supplier_link, price_per_unit)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                name.strip(), part_number.strip(), target, current,
                supplier_link.strip(), price
            ))

        conn.commit()
        conn.close()

        # ✅ Trigger success and next loop
        st.session_state.show_success = True
        st.session_state.show_next_action = False
        st.rerun()  # ✅ rerun instead of stop!


        # === Set Success Flags ===
        #st.session_state.show_success = True
        #st.rerun()



    # === SHOW DELETE CONFIRMATION ===
    # === SHOW DELETE CONFIRMATION THAT DISAPPEARS ===
    if st.session_state.product_deleted:
        st.success("🗑️ Product deleted successfully!")

        # Set the timestamp when message first appears
        if st.session_state.delete_timestamp is None:
            st.session_state.delete_timestamp = time.time()

        # Check if 3 seconds have passed since deletion
        elif time.time() - st.session_state.delete_timestamp >= 3:
            st.session_state.product_deleted = False
            st.session_state.delete_timestamp = None
            st.rerun()





    # === DELETE BUTTON ===
    if st.session_state.product_mode == "Edit" and selected_product:
        if st.button(f"❌ Delete '{selected_product}'"):
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM products WHERE name = ?", (selected_product,))
            conn.commit()
            conn.close()
            st.session_state.product_deleted = True
            st.session_state.clear_form = True
            st.rerun()














# === PAGE: MANAGE USERS ===
# === INITIAL STATE ===
# === INITIAL STATE ===
if "user_added" not in st.session_state:
    st.session_state.user_added = False
if "user_deleted" not in st.session_state:
    st.session_state.user_deleted = False
if "new_user_input" not in st.session_state:
    st.session_state.new_user_input = ""
if "reset_user_input" not in st.session_state:
    st.session_state.reset_user_input = False

# === PAGE: MANAGE USERS ===
if page == "Manage Users":
    st.title("👤 Manage Users")
    st.markdown("### ➕ Add or ❌ Delete a User")

    # ✅ Reset text field before rendering input
    if st.session_state.reset_user_input:
        st.session_state.new_user_input = ""
        st.session_state.reset_user_input = False

    # ✅ Show confirmation messages
    if st.session_state.user_added:
        st.success("✅ New user added successfully!")
        st.session_state.user_added = False

    if st.session_state.user_deleted:
        st.success("🗑️ User deleted successfully!")
        st.session_state.user_deleted = False

    # === FORM ===
    with st.form("user_form"):
        new_user = st.text_input("Add New User", key="new_user_input")
        user_to_delete = st.selectbox("Delete Existing User", [""] + get_users())
        submitted_user = st.form_submit_button("Confirm")

        if submitted_user:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            if new_user:
                cursor.execute("INSERT OR IGNORE INTO users (name) VALUES (?)", (new_user.strip(),))
                st.session_state.user_added = True
                st.session_state.reset_user_input = True  # ✅ Trigger clearing text

            if user_to_delete:
                cursor.execute("DELETE FROM users WHERE name = ?", (user_to_delete,))
                st.session_state.user_deleted = True  # ✅ Trigger delete message

            conn.commit()
            conn.close()

            st.rerun()















# === Utility Functions and App Setup ===
def get_users():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM users ORDER BY name")
    users = [row[0] for row in cursor.fetchall()]
    conn.close()
    return users

def get_products():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT name, part_number, target_qty, current_qty, supplier_link, price_per_unit FROM products ORDER BY name COLLATE NOCASE")
    products = cursor.fetchall()
    conn.close()
    return products


def update_inventory(product_name, qty_removed):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE products
        SET current_qty = current_qty - ?
        WHERE name = ?
    """, (qty_removed, product_name))
    conn.commit()
    conn.close()

def log_action(user_name, product_name, qty_removed):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("""
        INSERT INTO logs (user_name, product_name, qty_removed, timestamp)
        VALUES (?, ?, ?, ?)
    """, (user_name, product_name, qty_removed, timestamp))
    conn.commit()
    conn.close()










# === PAGE: Admin Upload ===
if page == "Admin Upload":
    st.title("🛠 Upload or Manage Product List")

    st.markdown("""
    Upload a **CSV or Excel file** to either *add to* or *completely replace* your current inventory list.

    **Required columns:** `PRODUCT`, `PRODUCT NUMBER`, `CURRENT QTY`, `TARGET QTY`, `SUPPLIER LINK`, `PRICE PER UNIT`
    """)

    # ⬇️ Download current inventory as backup
    if st.button("⬇️ Download Current Inventory as Backup"):
        conn = sqlite3.connect(DB_PATH)
        current_df = pd.read_sql_query("SELECT name, part_number, current_qty, target_qty, supplier_link, price_per_unit FROM products", conn)
        conn.close()

        csv_data = current_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download Inventory CSV",
            data=csv_data,
            file_name="current_inventory_backup.csv",
            mime="text/csv"
        )

    # 📤 Upload new inventory file
    uploaded_file = st.file_uploader("📤 Upload Inventory File (CSV or Excel)", type=["csv", "xlsx"])

    # --- SAMPLE FORMAT SECTION (View Only) ---
    st.markdown("### 🧾 Sample Format (Preview Only)")

    sample_df = pd.DataFrame({
        "PRODUCT": ["Example Product 1", "Example Product 2"],
        "PRODUCT NUMBER": ["ABC123", "XYZ456"],
        "CURRENT QTY": [100, 50],
        "TARGET QTY": [200, 75],
        "SUPPLIER LINK": ["https://supplier.com/abc", ""],
        "PRICE PER UNIT": [1.25, ""]
    })

    st.dataframe(sample_df, use_container_width=True)

    if uploaded_file:
        try:
            # Load file
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            df.dropna(how='all', inplace=True)
            preview_df = df.copy()  # Keep a copy for preview

            # Show uploaded file BEFORE renaming columns
            st.success("✅ File uploaded successfully!")
            # Show yellow reminder if changes haven't been applied yet
            if not st.session_state.get("upload_applied", False):
                st.markdown("""
                <div style="background-color: #fef9c3; padding: 10px; border-left: 6px solid #facc15; border-radius: 6px; margin-top: 10px;">
                ⚠️ <strong>Next Step:</strong> Choose how to apply the inventory update below and click <strong>'Apply Changes'</strong>.
                </div>
                """, unsafe_allow_html=True)




            if st.session_state.get("show_upload_success"):
                st.success("✅ Inventory successfully updated!")
                st.session_state.show_upload_success = False  # reset after showing

            st.dataframe(preview_df)

            # Rename columns for internal logic
            df.rename(columns={
                'PRODUCT': 'name',
                'PRODUCT NUMBER': 'part_number',
                'CURRENT QTY': 'current_qty',
                'TARGET QTY': 'target_qty',
                'SUPPLIER LINK': 'supplier_link',
                'PRICE PER UNIT': 'price_per_unit'
            }, inplace=True)

            # Confirm action
            action = st.radio("What would you like to do with this data?", ["Add to existing inventory", "Replace entire inventory"])

            if st.button("✅ Apply Changes"):
                required_cols = {'name', 'part_number', 'current_qty', 'target_qty', 'supplier_link', 'price_per_unit'}
                if not required_cols.issubset(df.columns):
                    st.error("❌ File must contain: 'PRODUCT', 'PRODUCT NUMBER', 'CURRENT QTY', 'TARGET QTY', 'SUPPLIER LINK', 'PRICE PER UNIT'")
                else:
                    # Format data
                    df['name'] = df['name'].astype(str).str.strip()
                    df['part_number'] = df['part_number'].astype(str).str.strip()
                    df['current_qty'] = pd.to_numeric(df['current_qty'], errors='coerce').fillna(0).astype(int)
                    df['target_qty'] = pd.to_numeric(df['target_qty'], errors='coerce').fillna(0).astype(int)
                    df['supplier_link'] = df['supplier_link'].astype(str).fillna("").str.strip()
                    df['price_per_unit'] = pd.to_numeric(df['price_per_unit'], errors='coerce')

                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()

                    if action == "Replace entire inventory":
                        cursor.execute("DELETE FROM products")
                        conn.commit()

                    for _, row in df.iterrows():
                        cursor.execute("""
                            INSERT OR REPLACE INTO products (name, part_number, current_qty, target_qty, supplier_link, price_per_unit)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            row['name'],
                            row['part_number'],
                            row['current_qty'],
                            row['target_qty'],
                            row.get('supplier_link', ''),
                            row.get('price_per_unit', None)
                        ))

                    conn.commit()
                    conn.close()

                    st.session_state.upload_applied = True  # ✅ Mark upload as applied
                    st.session_state.show_upload_success = True  # Trigger success bar on rerun
                    st.rerun()


        except Exception as e:
            st.error(f"❌ Something went wrong: {e}")
















# === PAGE: REPORTS ===
# === PAGE: REPORTS ===


import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
from io import BytesIO
import pandas as pd
import sqlite3


def get_logs():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM logs ORDER BY timestamp DESC", conn)
    conn.close()
    return df




def export_detailed_report_to_bytes(logs_df, db_path, selected_sheets=None, selected_months=None, selected_years=None):
    today = date.today()
    today_str = today.strftime('%m-%d-%Y')

    logs_df['Timestamp'] = pd.to_datetime(logs_df['timestamp'])
    logs_df['Date'] = logs_df['Timestamp'].dt.strftime('%m-%d-%Y')
    logs_today = logs_df[logs_df['Timestamp'].dt.date == today].copy()

    conn = sqlite3.connect(db_path)
    inventory_df = pd.read_sql_query("SELECT name as product_name, part_number, target_qty, current_qty FROM products", conn)
    restock_logs = pd.read_sql_query("SELECT * FROM restock_logs", conn)
    purchase_logs = pd.read_sql_query("SELECT * FROM purchase_logs", conn)

    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')






    # === USERS TOTAL ===
    if selected_sheets.get("Users Total"):
        logs_today_renamed = logs_today.rename(columns={
            'user_name': 'User',
            'product_name': 'Product Removed',
            'qty_removed': 'Qty Removed'
        })

        grouped_by_user = logs_today_renamed.groupby(['User', 'Product Removed']).agg({
            'Qty Removed': 'sum',
            'Date': 'min'
        }).reset_index()

        grouped_by_user = grouped_by_user.merge(
            inventory_df[['product_name', 'part_number', 'current_qty']],
            left_on='Product Removed',
            right_on='product_name',
            how='left'
        )

        grouped_by_user.drop(columns=['product_name'], inplace=True)
        grouped_by_user.rename(columns={'part_number': 'Part Number', 'current_qty': 'Current Stock'}, inplace=True)

        grouped_by_user = grouped_by_user[["User", "Product Removed", "Part Number", "Qty Removed", "Current Stock", "Date"]]
        grouped_by_user.to_excel(writer, sheet_name='Staff Usage Breakdown', index=False)

    # === PRODUCT TOTALS ===
    if selected_sheets.get("Product Totals"):
        product_totals = logs_today.groupby('product_name').agg({'qty_removed': 'sum'}).reset_index()
        product_totals = product_totals.merge(
            inventory_df,
            left_on='product_name',
            right_on='product_name',
            how='left'
        )
        product_totals.rename(columns={
            'product_name': 'Product Removed',
            'qty_removed': 'Qty Removed',
            'current_qty': 'Current Stock',
            'target_qty': '6 Month Target'
        }, inplace=True)
        product_totals['% Remaining'] = round((product_totals['Current Stock'] / product_totals['6 Month Target']) * 100, 1)
        product_totals['Date'] = today_str
        product_totals = product_totals[[
            'Product Removed',
            'part_number',
            'Qty Removed',
            'Current Stock',
            '6 Month Target',
            '% Remaining',
            'Date'
        ]]
        product_totals.to_excel(writer, sheet_name='Product Usage Summary', index=False)

    # === LOW STOCK ===
    if selected_sheets.get("Low Stock"):
        low_stock = inventory_df.copy()
        low_stock['% Remaining'] = round((low_stock['current_qty'] / low_stock['target_qty']) * 100, 1)
        low_stock = low_stock[low_stock['% Remaining'] < 65]
        low_stock.rename(columns={
            'product_name': 'Product Name',
            'current_qty': 'Current Stock',
            'target_qty': '6 Month Target'
        }, inplace=True)
        low_stock['Date'] = today_str
        low_stock = low_stock[['Product Name', 'Current Stock', '6 Month Target', '% Remaining', 'Date']]
        low_stock.to_excel(writer, sheet_name='Low Stock Alert', index=False)

    # === RESTOCK SUMMARY ===
    if selected_sheets.get("Restock Summary"):
        restock_logs['timestamp'] = pd.to_datetime(restock_logs['timestamp'])
        restock_logs['Date'] = restock_logs['timestamp'].dt.strftime('%m-%d-%Y')

        restock_summary = restock_logs.merge(
            inventory_df[['product_name', 'part_number', 'current_qty']],
            left_on='product',
            right_on='product_name',
            how='left'
        )
        restock_summary['Stock Before'] = restock_summary['current_qty'] - restock_summary['qty_added']
        restock_summary['Stock After'] = restock_summary['current_qty']
        restock_summary.rename(columns={
            'product': 'Product',
            'part_number': 'Part Number',
            'qty_added': 'Qty Restocked',
        }, inplace=True)
        restock_summary = restock_summary[['Date', 'Product', 'Part Number', 'Qty Restocked', 'Stock Before', 'Stock After']]
        restock_summary.to_excel(writer, sheet_name='Restock History', index=False)

    # === PURCHASE HISTORY ===
    if selected_sheets.get("Purchase History") and (selected_months or selected_years):
        purchase_logs = pd.read_sql_query("SELECT * FROM purchase_logs", conn)
        purchase_logs['timestamp'] = pd.to_datetime(purchase_logs['timestamp'])

        # Add Month and Year
        purchase_logs['Month'] = purchase_logs['timestamp'].dt.strftime('%B %Y')
        purchase_logs['Year'] = purchase_logs['timestamp'].dt.year

        # Filter by selected months
        filtered_logs = purchase_logs.copy()
        if selected_months:
            # ✅ Sort the selected months in calendar order
            from datetime import datetime
            selected_months.sort(key=lambda m: datetime.strptime(m, "%B %Y"))

            filtered_logs = filtered_logs[filtered_logs['Month'].isin(selected_months)]


        # Monthly Summary
        monthly_summary = []
        if selected_months:
            for month in selected_months:
                month_df = filtered_logs[filtered_logs['Month'] == month]

                # Make sure price_per_unit exists
                if 'price_per_unit' not in month_df.columns:
                    month_df['price_per_unit'] = 0.0
                # Make sure total_price exists
                if 'total_price' not in month_df.columns:
                    month_df['total_price'] = month_df['quantity'] * month_df['price_per_unit']
                
                grouped = month_df.groupby(['product_name', 'part_number']).agg(
                    Quantity=('quantity', 'sum'),
                    Price=('price_per_unit', 'mean'),
                    Total=('total_price', 'sum')
                ).reset_index()
                grouped.insert(0, 'Month', month)
                monthly_summary.append(grouped)
                monthly_summary.append(pd.DataFrame([[''] * grouped.shape[1]], columns=grouped.columns))  # blank row

        # Yearly Summary
        yearly_summary = []
        if selected_years:
            selected_years_int = [int(y) for y in selected_years]
            year_logs = purchase_logs[purchase_logs['Year'].isin(selected_years_int)]
            for year in selected_years_int:
                year_df = year_logs[year_logs['Year'] == year]

                # ✅ Fix: Use year_df instead of month_df
                if 'price_per_unit' not in year_df.columns:
                    year_df['price_per_unit'] = 0.0
                if 'total_price' not in year_df.columns:
                    year_df['total_price'] = year_df['quantity'] * year_df['price_per_unit']

                grouped = year_df.groupby(['product_name', 'part_number']).agg(
                    Quantity=('quantity', 'sum'),
                    Price=('price_per_unit', 'mean'),
                    Total=('total_price', 'sum')
                ).reset_index()
                grouped.insert(0, 'Month', f'Yearly Total {year}')
                yearly_summary.append(grouped)
                yearly_summary.append(pd.DataFrame([[''] * grouped.shape[1]], columns=grouped.columns))  # blank row

        # Combine summaries
        combined_df = pd.concat(monthly_summary + yearly_summary, ignore_index=True)

        # Rename columns
        combined_df.rename(columns={
            'product_name': 'Product',
            'part_number': 'Part Number',
            'Quantity': 'Quantity',
            'Price': 'Price Per Unit',
            'Total': 'Total Price'
        }, inplace=True)

        # Round price columns
        if 'Price Per Unit' in combined_df.columns:
            combined_df['Price Per Unit'] = combined_df['Price Per Unit'].round(2)
        if 'Total Price' in combined_df.columns:
            combined_df['Total Price'] = combined_df['Total Price'].round(2)

        # Reorder columns
        # Reorder columns
        combined_df = combined_df[['Month', 'Product', 'Part Number', 'Quantity', 'Price Per Unit', 'Total Price']]

        # Sort by Month to keep grouping consistent
        combined_df.sort_values(by="Month", inplace=True)















        # Insert blank rows between grouped months or years
        grouped_df = []
        first_group = True
        previous_month = None

        # Get all months/years the user selected, including ones not in the data
        all_labels = selected_months + [f"Yearly Total {y}" for y in selected_years]
        all_labels = list(dict.fromkeys(all_labels))  # remove duplicates, keep order

        for label in all_labels:
            # Get rows for this month/year
            group_rows = combined_df[combined_df['Month'] == label]

            if not first_group:
                # Add a blank row between sections
                blank_row = pd.Series([""] * len(combined_df.columns), index=combined_df.columns)
                grouped_df.append(blank_row)
            else:
                first_group = False

            if not group_rows.empty:
                first = True
                for _, row in group_rows.iterrows():
                    new_row = row.copy()
                    if not first:
                        new_row['Month'] = ""
                    else:
                        first = False
                    grouped_df.append(new_row)
            else:
                # If there are no purchases, add "No purchases for this month/year"
                blank_row = pd.Series([""] * len(combined_df.columns), index=combined_df.columns)
                blank_row['Month'] = label
                blank_row['Product'] = "No purchases for this month"
                grouped_df.append(blank_row)

        # Rebuild final DataFrame
        combined_df = pd.DataFrame(grouped_df, columns=combined_df.columns)

        # Remove blank rows only at the very beginning
        while len(combined_df) > 0 and all(str(cell).strip() == "" for cell in combined_df.iloc[0]):
            combined_df = combined_df.iloc[1:]




        # Export to Excel
        combined_df.to_excel(writer, sheet_name="Purchase History", index=False)

        # ✅ Apply dollar sign formatting to price columns in Excel
        workbook = writer.book
        worksheet = writer.sheets["Purchase History"]

        price_col_idx = combined_df.columns.get_loc("Price Per Unit") + 1
        total_col_idx = combined_df.columns.get_loc("Total Price") + 1

        currency_format = '$#,##0.00'

        for row in range(2, len(combined_df) + 2):  # start at row 2 (Excel rows start at 1, row 1 = headers)
            worksheet.cell(row=row, column=price_col_idx).number_format = currency_format
            worksheet.cell(row=row, column=total_col_idx).number_format = currency_format



    # ✅ Only close the writer if at least one sheet was written
    if writer.sheets:
        writer.close()
    else:
        st.warning("⚠️ No report was generated. Please select at least one sheet and required filters.")
        return None







    # ✅ Apply formatting
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(15, max_len + 2)
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = 'A2'
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output







    # Save raw report
    writer.close()
    output.seek(0)

    # === FORMAT WORKBOOK ===
    wb = openpyxl.load_workbook(output)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Format all cells + auto size
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
            col_letter = get_column_letter(col_cells[0].column)

            for cell in col_cells:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.column_dimensions[col_letter].width = max(15, min(max_len + 2, 40))  # min width 15, max 40

    # Save updated report
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()




# === STREAMLIT UI ===
if page == "Reports":
    st.title("📈 Inventory Usage Reports")

    logs_df = get_logs()
    if logs_df.empty:
        st.info("No usage logs found yet.")

    # === Show Daily Summary Table Instead of Full Logs ===
    today = date.today()
    logs_df['Timestamp'] = pd.to_datetime(logs_df['timestamp'])
    logs_df['Date'] = logs_df['Timestamp'].dt.date
    daily_summary = logs_df[logs_df['Date'] == today]

    if not daily_summary.empty:
        st.markdown("### Daily Summary (Today)")
        summary = (
            daily_summary.groupby(['user_name', 'product_name'])['qty_removed']
            .sum().reset_index().sort_values(by='qty_removed', ascending=False)
        )
        summary.columns = ['User', 'Product', 'Quantity Removed']
        st.dataframe(summary, use_container_width=True)
    else:
        st.info("No usage recorded today.")

    st.markdown("### Select Which Sheets to Include:")
    sheet_options = {
        "Users Total": st.checkbox("👤 Staff Usage Breakdwon", value=False),
        "Product Totals": st.checkbox("📦 Product Usage Summary", value=False),
        "Low Stock": st.checkbox("⚠️ Low Stock Alert", value=False),
        "Restock Summary": st.checkbox("🔄 Restock History", value=False),
        "Purchase History": st.checkbox("🛒 Purchase History", value=False),

    }



    # Only show month/year selection if Purchase History is selected
    if sheet_options.get("Purchase History"):

        conn = sqlite3.connect(DB_PATH)
        from calendar import month_name

        # 👇 Manually define full list of 2025 months
        full_months_2025 = [f"{month} 2025" for month in list(month_name)[1:]]  # Skip blank at index 0
        available_years = ["2025"]  # as strings to match your previous logic

        selected_months = st.multiselect("📅 Select Months for Purchase History", full_months_2025)
        selected_years = st.multiselect("📅 Select Years for Yearly Summary", available_years)


    else:
        selected_months = []
        selected_years = []




    

    if not any(sheet_options.values()):
        st.warning("⚠️ Please select at least one sheet.")
    else:
        report_bytes = export_detailed_report_to_bytes(
            logs_df,
            DB_PATH,
            selected_sheets=sheet_options,
            selected_months=selected_months,
            selected_years=selected_years
        )

        if report_bytes:
            st.download_button(
                label="📊 Generate and Download Excel Report",
                file_name="detailed_inventory_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                data=report_bytes,
                help="Click to generate and download the selected report as an Excel file."
            )




# === Reset dropdown flag after showing it once ===
if "show_supplier_link" in st.session_state:
    del st.session_state["show_supplier_link"]


# === FOOTER ===

st.markdown("""
<hr style="margin-top: 30px;"/>
<div style='text-align: center; font-size: 0.8rem; color: #999;'>
    © 2025 Nventory — Built for smart, simple inventory management.
</div>
""", unsafe_allow_html=True)



